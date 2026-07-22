import json
import csv
import sys
from pathlib import Path

CATEGORY_LABELS = {
    "mainstream": "Mainstream Legacy Media",
    "alternative": "Alternative / Right-wing Media",
    "social_media": "Social Media / Platforms",
    "state_funded": "State-funded / Affiliated",
    "institutional": "Institutional / Reference",
    "other": "Other",
}

def stats_to_rows(stats: dict) -> list[list]:
    rows = [["Domaine", "Catégorie", "Citations", "Sentiment moyen", "Écart-type", "Ratio négativité"]]
    for domain, s in sorted(stats.get("top_domains", {}).items(),
                            key=lambda x: -x[1]["count"]):
        rows.append([
            domain,
            "",  # category not stored per-domain in stats
            s["count"],
            round(s["mean_sentiment"], 4),
            round(s["std_sentiment"], 4),
            round(s["negativity_ratio"], 4),
        ])
    rows.append([])
    rows.append(["Catégorie", "Citations", "Sentiment moyen", "Écart-type", "Ratio négativité"])
    for cat_key, s in stats.get("categories", {}).items():
        rows.append([
            CATEGORY_LABELS.get(cat_key, cat_key),
            s["count"],
            round(s["mean_sentiment"], 4),
            round(s["std_sentiment"], 4),
            round(s["negativity_ratio"], 4),
        ])
    rows.append([])
    rows.append(["Total posts avec liens", stats.get("total_posts_with_links", 0)])
    rows.append(["Domaines uniques", stats.get("total_domains_seen", 0)])
    return rows


def write_xlsx(rows: list[list], posts_rows: list[list], path: str):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill
    except ImportError:
        print("openpyxl pas installé. Installation...")
        import subprocess
        subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
        import openpyxl
        from openpyxl.styles import Font, PatternFill

    wb = openpyxl.Workbook()

    # Sheet 1: Stats
    ws1 = wb.active
    ws1.title = "Stats consolidées"
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="333333", end_color="333333", fill_type="solid")
    for r_idx, row in enumerate(rows):
        for c_idx, val in enumerate(row):
            cell = ws1.cell(row=r_idx + 1, column=c_idx + 1, value=val)
            if r_idx == 0 or (len(row) > 1 and row[0] in CATEGORY_LABELS.values() and r_idx > len(stats_json.get("top_domains", {})) + 1):
                cell.font = header_font
                cell.fill = header_fill
    ws1.column_dimensions["A"].width = 35
    ws1.column_dimensions["B"].width = 30
    for c in "CDEF":
        ws1.column_dimensions[c].width = 18

    # Sheet 2: Posts individuels
    ws2 = wb.create_sheet("Posts")
    for r_idx, row in enumerate(posts_rows):
        for c_idx, val in enumerate(row):
            cell = ws2.cell(row=r_idx + 1, column=c_idx + 1, value=val)
            if r_idx == 0:
                cell.font = header_font
                cell.fill = header_fill
    ws2.column_dimensions["A"].width = 12
    ws2.column_dimensions["B"].width = 25
    ws2.column_dimensions["C"].width = 20
    ws2.column_dimensions["D"].width = 14
    ws2.column_dimensions["E"].width = 14
    ws2.column_dimensions["F"].width = 14
    ws2.column_dimensions["G"].width = 14
    ws2.column_dimensions["H"].width = 80

    wb.save(path)
    print(f"Fichier Excel créé : {path}")


def write_csv(rows: list[list], posts_rows: list[list], path: str):
    base = Path(path).with_suffix("")
    # Stats CSV
    with open(f"{base}_stats.csv", "w", newline="") as f:
        w = csv.writer(f)
        for row in rows:
            w.writerow(row)
    # Posts CSV
    with open(f"{base}_posts.csv", "w", newline="") as f:
        w = csv.writer(f)
        for row in posts_rows:
            w.writerow(row)
    print(f"Fichiers CSV créés : {base}_stats.csv et {base}_posts.csv")


if __name__ == "__main__":
    import argparse
    parser = argparse.ArgumentParser(description="Export consolidated stats sheet")
    parser.add_argument("stats_json", help="Stats JSON from pipeline")
    parser.add_argument("posts_csv", nargs="?", help="Posts CSV from pipeline (optional)")
    parser.add_argument("--output", "-o", default="consolidated_stats.xlsx", help="Output file")
    args = parser.parse_args()

    with open(args.stats_json) as f:
        stats_json = json.load(f)

    rows = stats_to_rows(stats_json)

    posts_rows = []
    if args.posts_csv:
        with open(args.posts_csv, newline="") as f:
            reader = csv.reader(f)
            posts_rows = list(reader)
    else:
        posts_rows = [["post_id", "primary_domain", "primary_category",
                       "compound", "neg_score", "neu_score", "pos_score", "text_preview"]]

    if args.output.endswith(".xlsx"):
        write_xlsx(rows, posts_rows, args.output)
    else:
        write_csv(rows, posts_rows, args.output)
